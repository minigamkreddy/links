import time
#from selenium.webdriver.support.ui import select
from selenium import webdriver
import sys
from time import gmtime,strftime
import openpyxl
import re
import logging
import socket
import httplib
import urllib
from selenium.webdriver.remote.command import Command

logging = logging.getLogger("__name__")

INPUT_XL_FILE = 'File.xlsx'
work_book = 0
sheet_names = 0
ap = 0
driver = 0
def gopi():
        ''' Initializes openpyxl and load the excel file.'''
        try:
	    global work_book
	    global sheet_names
            work_book = openpyxl.load_workbook(str(INPUT_XL_FILE))
	    sheet_names = work_book.sheetnames
        except IOError as  error:
            logging.error('%s file not found', INPUT_XL_FILE)
            raise Exception(error)
def _get_sheet_names():
        sheet_names = work_book.sheetnames
        return sheet_names

def _get_sheet_instance(sheet_name):
        return work_book[sheet_name]


def read_ap_information(index):
	global ap
        ap_info = get_ap_info(None)
	
        for ap in ap_info:
            if ap['Shipping'] == str(index):
                #if _input_validation(ap):
	        print ap
                return ap
        return False
	

def get_ap_info(index):
	sheet_info = get_excel_info('Sheet1', index)
        return sheet_info['Sheet1']

def get_excel_info(sheet_name, sheet_index):
        if sheet_name and not _validate_sheet_name(sheet_name):
            logging.error('{0} sheet in invalid'.format(sheet_name))
            logging.info('Sheets in current excel file: {0}' \
                       .format(_get_sheet_names()))
            return

	elif sheet_name and _validate_sheet_name(sheet_name):
	    return {sheet_name: _get_sheet_info( \
                  _get_sheet_instance(sheet_name), sheet_index)}
        # Get info for all the sheets
        # CHECK: Getting all the sheets values at once is that necessary ?

	else:
	    return False

def _get_sheet_info(sheet_instance, sheet_index):
        '''Gets the sheet information from excel file.

        Args:
          sheet(instance): Sheet instance
          sheet_index(int): Row number of the sheet

        Returns:
          Returns dictionary of particular row w.r.t sheet index, otherwise
          returns list of dictionaries which contains all sheets information.
        '''
        if sheet_instance.max_column == 2:
            return ({str(sheet_instance.cell(row, 1).value): \
                          str(sheet_instance.cell(row, 2).value) \
                          for row in xrange(1, sheet_instance.max_row+1)})

        if sheet_index:
            if _validate_sheet_index(sheet_instance, sheet_index):
                return filter(None, [{str(sheet_instance.cell(1, column).value):
                              str(sheet_instance.cell(sheet_index, column).value)
                              for column in range(1, sheet_instance.max_column+1)}])
            else:
                logging.error('Invalid sheet index (%s) for sheet : %s',
                              sheet_index, str(sheet_instance.title))
        else:
            return filter(None, [{str(sheet_instance.cell(1, column).value):
                            str(sheet_instance.cell(row, column).value)
                            for column in range(1, sheet_instance.max_column+1)}
                            for row in xrange(2, sheet_instance.max_row+1)])


def _validate_sheet_index(sheet_instance, sheet_index):
        '''Validates sheet's row number

        Args:
            sheet_index (int): Row number of the sheet

        Returns:
            True for success,otherwise False
        '''
        if sheet_index > 1 and sheet_instance.max_row+1 > sheet_index:
            return True
        else:
            return False
def _validate_sheet_name(sheet_name):
        '''Validates sheet's name.

        Args:
            sheet_name (str): Name of the sheet

        Returns:
            True for success,otherwise False
        '''
        if sheet_name is None or sheet_name in _get_sheet_names():
            return True
        else:
            return False


def _pre_req():
        driver=webdriver.Firefox()
#        driver.Set_Page_Load.timeout(180)
        driver.delete_all_cookies()
        return driver
def Login(http_addr):
        try:
	  global driver
	  driver = _pre_req()
          logging.info("http://"+http_addr)
          driver.get("http://" + http_addr)
          time.sleep(4)  # AP login page loading time
          #self.driver.find_element_by_xpath('//*[@id="login-username"'
          #                                  ']').send_keys(user_name)
	  print(ap['IEC Code'])
	  driver.find_element_by_xpath('/html/body/form/div/center/table/tbody/tr[1]/td[2]/input').send_keys(ap['IEC Code'])
	  time.sleep(4)
	  driver.find_element_by_xpath('/html/body/form/div/center/table/tbody/tr[1]/td[4]/input').send_keys(ap['IFSC Code'])
	  time.sleep(4)
	  img = driver.find_element_by_xpath('/html/body/form/div/center/table/tbody/tr[6]/td[3]/img')
          src = img.get_attribute('src')
	  urllib.request.urlretrieve(src, "captcha.jpeg")
	  print(src)
	 
          time.sleep(4)
	  driver.find_element_by_xpath('/html/body/form/div/center/table/tbody/tr[6]/td[4]/font').send_keys(src)
	  driver.find_element_by_xpath('/html/body/form/div/center/table/tbody/tr[7]/td/p/input[1]').click()
          #time.sleep(4)  # AP page loading time
	except Exception as error_msg:
          #logging.error('Exception in AP login : %s' % error_msg)
          return False
        return True


def CloseWebDriver():
        """ Closes the webdriver

        Args:
            None

        Returns:
            bool: retVal. True for success, False otherwise.
        """
        driver.quit()
        try:
            driver.execute(Command.STATUS)
            logging.error('Webdriver not closed')
            return False
        except (socket.error, httplib.CannotSendRequest):
            return True
        return True
gopi()
ap = read_ap_information('3678118')
print(ap)
Login('dgftebrc.nic.in:8100/BRCQueryTrade/index.jsp')
CloseWebDriver()
